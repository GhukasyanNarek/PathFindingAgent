import random
import time
import pygame
from game import Game
from algorithms import Strategy
from openpyxl import Workbook, load_workbook
import copy

WIDTH = 800
WIN = pygame.display.set_mode((WIDTH, WIDTH))
pygame.display.set_caption("Path Finding Agent")

ROWS = 50
OBSTACLE_DENSITY = 0.3


def generate_random_points(rows, grid):
    """
    Generates random start and end positions ensuring they are not on barriers.

    Args:
        rows (int): Number of rows in the grid.
        grid (list): The grid containing Spot objects.

    Returns:
        tuple: Start and end positions as (row, col) tuples.
    """
    while True:
        start = (random.randint(0, rows - 1), random.randint(0, rows - 1))
        end = (random.randint(0, rows - 1), random.randint(0, rows - 1))
        if (start != end and
            not grid[start[0]][start[1]].is_barrier() and
            not grid[end[0]][end[1]].is_barrier()):
            return start, end


def place_obstacles(grid, density, start_pos, end_pos):
    """
    Places obstacles on the grid based on the specified density, avoiding start and end positions.

    Args:
        grid (list): The grid containing Spot objects.
        density (float): Obstacle density (0 to 1).
        start_pos (tuple): Start position as (row, col).
        end_pos (tuple): End position as (row, col).
    """
    total_cells = ROWS * ROWS
    obstacle_count = int(total_cells * density)
    obstacles = set()
    while len(obstacles) < obstacle_count:
        row, col = random.randint(0, ROWS - 1), random.randint(0, ROWS - 1)
        if ((row, col) not in obstacles and
            (row, col) != start_pos and
            (row, col) != end_pos and
            not grid[row][col].is_barrier()):
            obstacles.add((row, col))
            grid[row][col].make_barrier()


def save_metrics_to_xlsx(metrics, filename="data.xlsx"):
    """
    Saves the pathfinding metrics to an Excel (.xlsx) file.

    Args:
        metrics (dict): The dictionary of metrics to save.
        filename (str): The name of the Excel file.
    """
    try:
        # Load the workbook if it exists; otherwise, create a new one
        try:
            workbook = load_workbook(filename)
            sheet = workbook.active
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active
            # Add header row
            sheet.append(["Path", "Execution Time (s)", "Steps", "Manhattan Distance",
                         "Expanded Nodes", "Algorithm", "Mode", "Run"])

        # Append the metrics to the worksheet
        sheet.append([
            str(metrics["path"]),            # Path as a string
            metrics["time"],                 # Execution time in seconds
            metrics["steps"],                # Number of steps in the path
            metrics["manhattan_distance"],   # Manhattan distance
            metrics["expanded_nodes"],       # Number of expanded nodes
            metrics["algorithm"],            # Algorithm type
            metrics["mode"],                 # Mode used
        ])

        # Save the workbook
        workbook.save(filename)
        print(f"Metrics saved to {filename}")
    except Exception as e:
        print(f"An error occurred while saving to {filename}: {e}")


def automated_tests(win, width, algorithm, num_tests=100, mode='2'):
    """
    Runs automated tests for a single algorithm on different random grids for each test.

    Args:
        win (pygame.Surface): The pygame window surface.
        width (int): Width of the window.
        algorithm (function): The pathfinding algorithm to test.
        num_tests (int): Number of tests to run.
        mode (str): Mode identifier.
    """
    for run_id in range(1, num_tests + 1):
        print(f"Running test {run_id} on a new random grid...")

        # Create a new grid for each test
        grid = Game.make_grid(ROWS, width)
        start_pos, end_pos = generate_random_points(ROWS, grid)

        # Initialize start, end, and obstacles on the grid
        start = grid[start_pos[0]][start_pos[1]]
        end = grid[end_pos[0]][end_pos[1]]
        start.make_start()
        end.make_end()
        place_obstacles(grid, OBSTACLE_DENSITY, start_pos, end_pos)

        # Update neighbors for all spots in the grid
        for row in grid:
            for spot in row:
                spot.update_neighbors(grid)

        # Run the algorithm
        metrics = algorithm(lambda: Game.draw(win, grid, ROWS, width), grid, start, end)
        if metrics:
            metrics["mode"] = mode
            metrics["run"] = run_id
            save_metrics_to_xlsx(metrics)
        else:
            print(f"Test {run_id}: No path found.")

    print("All tests completed.")


def run_all_algorithms_for_configurations(win, width, num_tests=100, mode='3'):
    """
    Runs automated tests for multiple algorithms on different random grids for each test.

    Args:
        win (pygame.Surface): The pygame window surface.
        width (int): Width of the window.
        num_tests (int): Number of tests to run.
        mode (str): Mode identifier.
    """
    algorithms = ["a_star", "bfs", "dfs", "greedy_bfs"]

    for test_number in range(1, num_tests + 1):
        print(f"Running test {test_number} on a new random grid...")

        # Create a new grid for this test
        grid = Game.make_grid(ROWS, width)
        start_pos, end_pos = generate_random_points(ROWS, grid)

        # Initialize start, end, and obstacles on the grid
        start = grid[start_pos[0]][start_pos[1]]
        end = grid[end_pos[0]][end_pos[1]]
        start.make_start()
        end.make_end()
        place_obstacles(grid, OBSTACLE_DENSITY, start_pos, end_pos)

        # Run each algorithm on the grid
        for algo_name in algorithms:
            try:
                print(f"Running {algo_name} on test {test_number}...")

                # Assign algorithm function
                func = getattr(Strategy, algo_name)

                # Deep copy the grid for the current algorithm
                algo_grid = copy.deepcopy(grid)

                # Reassign start and end in the copied grid
                grid_start = algo_grid[start_pos[0]][start_pos[1]]
                grid_end = algo_grid[end_pos[0]][end_pos[1]]
                grid_start.make_start()
                grid_end.make_end()

                # Update neighbors on the copied grid
                for row in algo_grid:
                    for spot in row:
                        spot.update_neighbors(algo_grid)

                # Measure execution time
                start_time = time.time()
                metrics = func(lambda: Game.draw(win, algo_grid, ROWS, width), algo_grid, grid_start, grid_end)
                exec_time = time.time() - start_time

                if metrics:
                    metrics.update({
                        "run": test_number,
                        "algorithm": algo_name,
                        "time": exec_time,
                        "mode": mode
                    })
                    save_metrics_to_xlsx(metrics)
                else:
                    print(f"Test {test_number}, Algorithm {algo_name}: No path found.")
            except Exception as e:
                print(f"Test {test_number}, Algorithm {algo_name}: Error occurred - {e}")

        print(f"All algorithms completed for test {test_number}. Moving to the next grid...")

    print("All tests completed.")


def main(win, width, algorithm=Strategy.a_star, mode='1'):
    """
    Runs the main manual mode where users can interactively place obstacles, start, and end points.

    Args:
        win (pygame.Surface): The pygame window surface.
        width (int): Width of the window.
        algorithm (function): The pathfinding algorithm to execute.
        mode (str): Mode identifier.
    """
    grid = Game.make_grid(ROWS, width)
    start = None
    end = None
    run_flag = True
    while run_flag:
        Game.draw(win, grid, ROWS, width)
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                run_flag = False

            if pygame.mouse.get_pressed()[0]:  # Left mouse button
                pos = pygame.mouse.get_pos()
                row, col = Game.get_clicked_pos(pos, ROWS, width)
                spot = grid[row][col]
                if not start and spot != end:
                    start = spot
                    start.make_start()
                elif not end and spot != start:
                    end = spot
                    end.make_end()
                elif spot != end and spot != start:
                    spot.make_barrier()

            elif pygame.mouse.get_pressed()[2]:  # Right mouse button
                pos = pygame.mouse.get_pos()
                row, col = Game.get_clicked_pos(pos, ROWS, width)
                spot = grid[row][col]
                spot.reset()
                if spot == start:
                    start = None
                elif spot == end:
                    end = None

            if event.type == pygame.KEYDOWN:
                if event.key == pygame.K_SPACE and start and end:
                    for row in grid:
                        for spot in row:
                            spot.update_neighbors(grid)

                    metrics = algorithm(lambda: Game.draw(win, grid, ROWS, width), grid, start, end)

                    if metrics:
                        metrics['mode'] = mode
                        print("Pathfinding Metrics:")
                        print(metrics)
                        save_metrics_to_xlsx(metrics, "data.xlsx")
                    else:
                        print("No path found!")

                if event.key == pygame.K_c:
                    start = None
                    end = None
                    grid = Game.make_grid(ROWS, width)

    pygame.quit()


if __name__ == "__main__":
    choice = input("Enter '1' for manual mode, '2' for automated tests, or '3' for automated multi-algorithm tests: ").strip()
    mode = '1' if choice == '1' else ('2' if choice == '2' else '3')
    pygame.init()
    try:
        if choice == "1":
            main(WIN, WIDTH, Strategy.greedy_bfs, mode=mode)
        elif choice == "2":
            automated_tests(WIN, WIDTH, Strategy.a_star, num_tests=5, mode=mode)
        elif choice == "3":
            run_all_algorithms_for_configurations(WIN, WIDTH, num_tests=2, mode=mode)
        else:
            print("Invalid choice. Exiting.")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
    finally:
        pygame.quit()